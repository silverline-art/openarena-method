"""Group-level statistical comparison and visualization (v1.4.0)"""
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import logging
from scipy import stats
import openpyxl

logger = logging.getLogger(__name__)


class GroupComparator:
    """
    Perform group-level statistical comparisons and generate publication-quality plots.

    Compares metrics across experimental groups (control, grade0.1, grade1, grade5)
    with median/deviation bars and statistical significance testing.
    """

    def __init__(self, output_dir: Path):
        """
        Initialize group comparator.

        Args:
            output_dir: Base output directory containing group subdirectories
        """
        self.output_dir = Path(output_dir)
        self.group_data = {}

    def load_group_data(self, group_mapping: Dict[str, Dict]) -> None:
        """
        Load all Excel files for each group and extract metrics.

        Args:
            group_mapping: Dictionary mapping group names to sample configurations
                          e.g., {'control': {'samples': ['control-1', ...]}, ...}
        """
        logger.info("Loading data from all groups...")

        for group_name, group_info in group_mapping.items():
            logger.info(f"  Loading group: {group_name}")
            group_samples = []

            # Map group name to directory name
            group_dir_map = {
                'control': 'control',
                'low_dose_01': 'low_dose_01',
                'medium_dose_1': 'medium_dose_1',
                'high_dose_5': 'high_dose_5'
            }

            group_dir = self.output_dir / group_dir_map.get(group_name, group_name)

            if not group_dir.exists():
                logger.warning(f"    Group directory not found: {group_dir}")
                continue

            # Find all sample Excel files
            for sample_id in group_info['samples']:
                sample_dir = group_dir / sample_id
                if not sample_dir.exists():
                    logger.warning(f"    Sample directory not found: {sample_dir}")
                    continue

                # Find most recent Excel file
                excel_files = list(sample_dir.glob("Gait_Analysis_*.xlsx"))
                if not excel_files:
                    logger.warning(f"    No Excel file found for sample: {sample_id}")
                    continue

                excel_file = sorted(excel_files)[-1]  # Most recent

                try:
                    sample_metrics = self._load_sample_metrics(excel_file)
                    sample_metrics['sample_id'] = sample_id
                    group_samples.append(sample_metrics)
                except Exception as e:
                    logger.error(f"    Failed to load {excel_file}: {e}")

            self.group_data[group_name] = group_samples
            logger.info(f"    Loaded {len(group_samples)} samples for {group_name}")

    def _load_sample_metrics(self, excel_path: Path) -> Dict:
        """
        Load metrics from a single Excel file.

        Args:
            excel_path: Path to Excel file

        Returns:
            Dictionary containing all metrics for this sample
        """
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        metrics = {}

        # Load Gait Metrics sheet
        if 'Gait Metrics' in wb.sheetnames:
            ws = wb['Gait Metrics']
            df = pd.DataFrame(ws.values)
            df.columns = df.iloc[0]
            df = df[1:]  # Remove header row

            # Extract metrics with 'value' field (aggregated metrics)
            for _, row in df.iterrows():
                limb = row['limb']
                metric = row['metric']
                value = row['value']

                if pd.notna(value) and metric not in ['num_strides']:
                    key = f"{limb}_{metric}"
                    try:
                        metrics[key] = float(value)
                    except (ValueError, TypeError):
                        pass

        # Load ROM Metrics sheet
        if 'ROM Metrics' in wb.sheetnames:
            ws = wb['ROM Metrics']
            df = pd.DataFrame(ws.values)
            df.columns = df.iloc[0]
            df = df[1:]

            for _, row in df.iterrows():
                joint = row['joint']
                metric = row['metric']
                value = row['value']

                if pd.notna(value):
                    key = f"{joint}_{metric}"
                    try:
                        metrics[key] = float(value)
                    except (ValueError, TypeError):
                        pass

        wb.close()
        return metrics

    def compute_group_statistics(self) -> pd.DataFrame:
        """
        Compute median, mean, std for each metric across groups.

        Returns:
            DataFrame with group statistics
        """
        logger.info("Computing group statistics...")

        stats_rows = []

        # Get all unique metrics across all groups
        all_metrics = set()
        for group_samples in self.group_data.values():
            for sample in group_samples:
                all_metrics.update(sample.keys())
        all_metrics.discard('sample_id')

        for metric in sorted(all_metrics):
            row = {'metric': metric}

            for group_name in ['control', 'low_dose_01', 'medium_dose_1', 'high_dose_5']:
                if group_name not in self.group_data:
                    continue

                # Extract values for this metric from all samples in group
                values = []
                for sample in self.group_data[group_name]:
                    if metric in sample and pd.notna(sample[metric]):
                        values.append(sample[metric])

                if len(values) > 0:
                    row[f'{group_name}_median'] = np.median(values)
                    row[f'{group_name}_mean'] = np.mean(values)
                    row[f'{group_name}_std'] = np.std(values)
                    # Median Absolute Deviation (MAD) - robust measure of variability
                    row[f'{group_name}_mad'] = np.median(np.abs(values - np.median(values)))
                    row[f'{group_name}_n'] = len(values)
                    row[f'{group_name}_values'] = values  # Store raw values for plotting
                else:
                    row[f'{group_name}_median'] = np.nan
                    row[f'{group_name}_mean'] = np.nan
                    row[f'{group_name}_std'] = np.nan
                    row[f'{group_name}_mad'] = np.nan
                    row[f'{group_name}_n'] = 0
                    row[f'{group_name}_values'] = []

            stats_rows.append(row)

        df = pd.DataFrame(stats_rows)
        logger.info(f"  Computed statistics for {len(df)} metrics")
        return df

    def compute_statistical_significance(self, stats_df: pd.DataFrame) -> pd.DataFrame:
        """
        Compute p-values comparing each group to control using Mann-Whitney U test.

        Args:
            stats_df: DataFrame with group statistics

        Returns:
            Updated DataFrame with p-values
        """
        logger.info("Computing statistical significance (Mann-Whitney U test)...")

        for idx, row in stats_df.iterrows():
            metric = row['metric']
            control_values = row.get('control_values', [])

            if len(control_values) < 2:
                continue

            # Compare each treatment group to control
            for group_name in ['low_dose_01', 'medium_dose_1', 'high_dose_5']:
                group_values = row.get(f'{group_name}_values', [])

                if len(group_values) < 2:
                    stats_df.at[idx, f'{group_name}_pvalue'] = np.nan
                    continue

                # Mann-Whitney U test (non-parametric, no normality assumption)
                try:
                    statistic, pvalue = stats.mannwhitneyu(
                        control_values,
                        group_values,
                        alternative='two-sided'
                    )
                    stats_df.at[idx, f'{group_name}_pvalue'] = pvalue
                except Exception as e:
                    logger.warning(f"    Statistical test failed for {metric}, {group_name}: {e}")
                    stats_df.at[idx, f'{group_name}_pvalue'] = np.nan

        return stats_df

    def create_group_comparison_plots(self, stats_df: pd.DataFrame, output_dir: Path) -> None:
        """
        Create publication-quality group comparison plots for all metrics.

        Args:
            stats_df: DataFrame with group statistics and p-values
            output_dir: Directory to save plots
        """
        logger.info("Creating group comparison plots...")

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # Group metrics by category
        gait_metrics = [m for m in stats_df['metric'] if any(x in m for x in ['cadence', 'duty_cycle', 'speed', 'stride', 'swing_stance', 'regularity', 'phase_dispersion'])]
        rom_metrics = [m for m in stats_df['metric'] if any(x in m for x in ['rom', 'angular', 'sway'])]

        # Create plots for gait metrics
        self._create_metric_group_plots(
            stats_df[stats_df['metric'].isin(gait_metrics)],
            output_dir / 'group_comparison_gait_metrics.png',
            title='Gait Metrics - Group Comparison'
        )

        # Create plots for ROM metrics
        self._create_metric_group_plots(
            stats_df[stats_df['metric'].isin(rom_metrics)],
            output_dir / 'group_comparison_rom_metrics.png',
            title='Range of Motion Metrics - Group Comparison'
        )

        # Create individual plots for ALL metrics
        logger.info(f"Creating individual plots for all {len(stats_df)} metrics...")
        for _, metric_row in stats_df.iterrows():
            metric_name = metric_row['metric']
            self._create_single_metric_plot(
                metric_row,
                output_dir / f'group_comparison_{metric_name}.png'
            )

        logger.info(f"  Saved plots to {output_dir}")

    def _create_metric_group_plots(self, metrics_df: pd.DataFrame, output_path: Path, title: str) -> None:
        """Create multi-panel plot for a group of metrics"""
        n_metrics = len(metrics_df)
        if n_metrics == 0:
            return

        # Calculate grid dimensions
        n_cols = min(4, n_metrics)
        n_rows = (n_metrics + n_cols - 1) // n_cols

        fig, axes = plt.subplots(n_rows, n_cols, figsize=(5*n_cols, 4*n_rows))
        fig.suptitle(title, fontsize=16, fontweight='bold')

        if n_rows == 1 and n_cols == 1:
            axes = np.array([[axes]])
        elif n_rows == 1:
            axes = axes.reshape(1, -1)
        elif n_cols == 1:
            axes = axes.reshape(-1, 1)

        for idx, (_, row) in enumerate(metrics_df.iterrows()):
            ax_row = idx // n_cols
            ax_col = idx % n_cols
            ax = axes[ax_row, ax_col]

            self._plot_metric_comparison(row, ax)

        # Hide empty subplots
        for idx in range(n_metrics, n_rows * n_cols):
            ax_row = idx // n_cols
            ax_col = idx % n_cols
            axes[ax_row, ax_col].axis('off')

        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()

    def _create_single_metric_plot(self, metric_row: pd.Series, output_path: Path) -> None:
        """Create a single high-quality plot for one metric"""
        fig, ax = plt.subplots(figsize=(8, 6))
        self._plot_metric_comparison(metric_row, ax)
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()

    def _plot_metric_comparison(self, metric_row: pd.Series, ax) -> None:
        """
        Plot comparison for a single metric across groups.

        Shows median bars with MAD (median absolute deviation) error bars and p-value annotations.
        """
        metric_name = metric_row['metric']

        groups = ['control', 'low_dose_01', 'medium_dose_1', 'high_dose_5']
        group_labels = ['Control\n(0 Gy)', 'Grade 0.1\n(0.1 Gy)', 'Grade 1\n(1 Gy)', 'Grade 5\n(5 Gy)']
        colors = ['#2E7D32', '#FFA726', '#EF5350', '#7E57C2']

        medians = []
        mads = []
        x_pos = []

        for i, group in enumerate(groups):
            median = metric_row.get(f'{group}_median', np.nan)
            mad = metric_row.get(f'{group}_mad', np.nan)

            if pd.notna(median):
                medians.append(median)
                mads.append(mad if pd.notna(mad) else 0)
                x_pos.append(i)

        if len(medians) == 0:
            ax.text(0.5, 0.5, 'No data available', ha='center', va='center', transform=ax.transAxes)
            ax.set_title(metric_name.replace('_', ' ').title(), fontsize=10)
            return

        # Plot bars with MAD error bars
        bars = ax.bar(x_pos, medians, yerr=mads, color=[colors[i] for i in x_pos],
                     alpha=0.7, edgecolor='black', linewidth=1.5,
                     error_kw={'linewidth': 2, 'ecolor': 'black', 'capsize': 5})

        # Add individual data points as scatter
        for i, group in enumerate(groups):
            values = metric_row.get(f'{group}_values', [])
            if len(values) > 0:
                x_scatter = np.full(len(values), i) + np.random.normal(0, 0.05, len(values))
                ax.scatter(x_scatter, values, color='black', alpha=0.4, s=30,
                          zorder=3, edgecolors='white', linewidth=0.5)

        # Add p-value annotations
        y_max = max(medians) + max(mads) if len(mads) > 0 else max(medians)
        y_range = y_max - min(medians) if len(medians) > 1 else y_max

        for i, group in enumerate(['low_dose_01', 'medium_dose_1', 'high_dose_5']):
            pvalue = metric_row.get(f'{group}_pvalue', np.nan)
            if pd.notna(pvalue):
                # Add significance stars
                if pvalue < 0.001:
                    sig_text = '***'
                elif pvalue < 0.01:
                    sig_text = '**'
                elif pvalue < 0.05:
                    sig_text = '*'
                else:
                    sig_text = 'ns'

                # Draw significance bracket
                x1, x2 = 0, i + 1
                y = y_max + 0.1 * y_range + (i * 0.05 * y_range)
                h = 0.02 * y_range

                ax.plot([x1, x1, x2, x2], [y, y+h, y+h, y], lw=1.5, c='black')
                ax.text((x1+x2)/2, y+h, sig_text, ha='center', va='bottom', fontsize=9, fontweight='bold')

        # Formatting
        ax.set_xticks(range(len(group_labels)))
        ax.set_xticklabels([group_labels[i] for i in x_pos], fontsize=9)
        ax.set_ylabel('Value', fontsize=10, fontweight='bold')
        ax.set_title(metric_name.replace('_', ' ').title(), fontsize=11, fontweight='bold')
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    def export_statistics_table(self, stats_df: pd.DataFrame, output_path: Path) -> None:
        """
        Export group statistics to Excel file.

        Args:
            stats_df: DataFrame with group statistics
            output_path: Path to output Excel file
        """
        logger.info(f"Exporting statistics table to {output_path}")

        # Create clean export DataFrame
        export_df = stats_df.copy()

        # Drop raw values columns
        value_cols = [col for col in export_df.columns if col.endswith('_values')]
        export_df = export_df.drop(columns=value_cols)

        # Round numeric columns
        numeric_cols = export_df.select_dtypes(include=[np.number]).columns
        export_df[numeric_cols] = export_df[numeric_cols].round(4)

        # Save to Excel
        export_df.to_excel(output_path, index=False, sheet_name='Group Statistics')
        logger.info(f"  Exported {len(export_df)} rows to {output_path}")
